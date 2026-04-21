"""
Management command to import compounds from ELN-linked Excel spreadsheets.

This command imports compounds from spreadsheets that use ELN (Electronic Lab Notebook)
references with hyperlinks. It extracts:
- ELN reference labels (e.g., 'KF001') and URLs from hyperlinked cells
- Chemdraw file links
- LCMS/HRMS QC document links
- HELM notation and sequence display strings

Usage:
    ccp4-python manage.py import_eln_spreadsheet \\
        --file /path/to/spreadsheet.xlsx \\
        --target "Target Name" \\
        --dry-run

    ccp4-python manage.py import_eln_spreadsheet \\
        --file /path/to/spreadsheet.xlsx \\
        --target "Target Name" \\
        --create-suppliers \\
        --verbose
"""

import re
from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None


class Command(BaseCommand):
    """Import compounds from ELN-linked Excel spreadsheets."""

    help = "Import compounds from Excel spreadsheets with ELN hyperlinks"

    # Column name mappings (case-insensitive)
    COLUMN_MAPPINGS = {
        'ncl_id': ['ncl_id', 'compound_id', 'id'],
        'batch': ['batch', 'batch_number'],
        'peptide_id': ['peptide_id', 'peptide id', 'compound_name', 'name'],
        'target': ['target', 'project'],
        'eln_reference': ['eln_reference', 'eln reference', 'eln', 'notebook'],
        'chemdraw_file': ['chemdraw_file', 'chemdraw file', 'chemdraw', 'structure'],
        'chemical_formula': ['chemical_formula', 'chemical formula', 'formula'],
        'exact_mass': ['exact_mass', 'exact mass', 'mass', 'mw', 'molecular_weight'],
        'sequence': ['sequence', 'sequence_display'],
        'helm': ['helm', 'helm_notation'],
        'smiles': ['smiles'],
        'lcms': ['lcms', 'lc-ms', 'lcms_file'],
        'hrms': ['hrms', 'hr-ms', 'hrms_file'],
        'purity': ['purity', 'purity(%)', 'purity (%)'],
        'synthesized_by': ['synthesized_by', 'synthesied by', 'synthesized by', 'chemist'],
    }

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to the Excel spreadsheet (.xlsx)',
        )
        parser.add_argument(
            '--target',
            type=str,
            required=True,
            help='Target name to assign compounds to (must already exist)',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            help='Sheet name to import (default: first/active sheet)',
        )
        parser.add_argument(
            '--create-suppliers',
            action='store_true',
            help='Auto-create suppliers from ELN reference initials (e.g., KF -> Kallie Friston)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Parse and validate without saving to database',
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Show detailed progress',
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            help='Skip rows where compound already exists (by SMILES match)',
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl is required. Install with: pip install openpyxl")

        file_path = Path(options['file'])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        self.verbose = options['verbose']
        self.dry_run = options['dry_run']
        self.create_suppliers = options['create_suppliers']
        self.skip_existing = options['skip_existing']

        # Import models here to avoid issues during Django setup
        from compounds.registry.models import (
            Supplier, Target, Compound, Batch, LabNotebookEntry, CompoundDocument
        )

        # Validate target exists
        target_name = options['target']
        try:
            target = Target.objects.get(name=target_name)
        except Target.DoesNotExist:
            raise CommandError(f"Target not found: {target_name}")

        self.stdout.write(f"\nImporting from: {file_path}")
        self.stdout.write(f"Target: {target.name}")
        if self.dry_run:
            self.stdout.write(self.style.WARNING("[DRY RUN - no changes will be saved]"))
        self.stdout.write("-" * 60)

        # Load workbook
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
        except Exception as e:
            raise CommandError(f"Failed to open workbook: {e}")

        # Select sheet
        if options['sheet']:
            if options['sheet'] not in wb.sheetnames:
                raise CommandError(f"Sheet not found: {options['sheet']}. Available: {wb.sheetnames}")
            ws = wb[options['sheet']]
        else:
            ws = wb.active

        self.log(f"Sheet: {ws.title}")
        self.log(f"Dimensions: {ws.dimensions}")

        # Parse header row
        header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        column_map = self._map_columns(header_row)
        self.log(f"Mapped columns: {list(column_map.keys())}")

        # Check required columns
        required = ['smiles']  # At minimum we need SMILES to register
        missing = [col for col in required if col not in column_map]
        if missing:
            raise CommandError(f"Missing required columns: {missing}")

        # Track statistics
        stats = defaultdict(int)
        suppliers_created = []
        notebook_entries_created = []
        compounds_created = []
        documents_created = []
        errors = []

        # Process data rows
        rows = list(ws.iter_rows(min_row=2))
        self.stdout.write(f"Processing {len(rows)} rows...")

        with transaction.atomic():
            for row_idx, row in enumerate(rows, start=2):
                try:
                    result = self._process_row(
                        row, row_idx, column_map, target,
                        Supplier, Compound, Batch, LabNotebookEntry, CompoundDocument,
                        stats, suppliers_created, notebook_entries_created,
                        compounds_created, documents_created
                    )
                    if result:
                        stats['rows_processed'] += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {e}")
                    stats['errors'] += 1
                    if self.verbose:
                        self.stdout.write(self.style.ERROR(f"  Row {row_idx}: {e}"))

            if self.dry_run:
                # Rollback transaction in dry-run mode
                transaction.set_rollback(True)

        # Print summary
        self.stdout.write("\n" + "=" * 60)
        self.stdout.write("SUMMARY")
        self.stdout.write("=" * 60)
        self.stdout.write(f"Rows processed: {stats['rows_processed']}")
        self.stdout.write(f"Suppliers created: {len(suppliers_created)}")
        for s in suppliers_created:
            self.stdout.write(f"  - {s}")
        self.stdout.write(f"Notebook entries created: {len(notebook_entries_created)}")
        self.stdout.write(f"Compounds created: {len(compounds_created)}")
        self.stdout.write(f"Documents created: {len(documents_created)}")
        self.stdout.write(f"Rows skipped (existing): {stats['skipped_existing']}")
        self.stdout.write(f"Rows skipped (empty SMILES): {stats['skipped_empty']}")
        self.stdout.write(f"Errors: {stats['errors']}")

        if errors:
            self.stdout.write(self.style.ERROR("\nErrors:"))
            for err in errors[:20]:  # Limit to first 20
                self.stdout.write(f"  {err}")
            if len(errors) > 20:
                self.stdout.write(f"  ... and {len(errors) - 20} more")

        if self.dry_run:
            self.stdout.write(self.style.WARNING("\n[DRY RUN - no changes were saved]"))
        else:
            self.stdout.write(self.style.SUCCESS("\nImport completed successfully!"))

    def log(self, message):
        """Log message if verbose mode is enabled."""
        if self.verbose:
            self.stdout.write(f"  {message}")

    def _map_columns(self, header_row):
        """Map header row to known column names."""
        column_map = {}
        for col_idx, cell in enumerate(header_row):
            if cell.value is None:
                continue
            header = str(cell.value).lower().strip()
            for canonical, aliases in self.COLUMN_MAPPINGS.items():
                if header in aliases:
                    column_map[canonical] = col_idx
                    break
        return column_map

    def _get_cell_with_hyperlink(self, row, col_idx):
        """Extract cell value and hyperlink target."""
        if col_idx is None or col_idx >= len(row):
            return None, None
        cell = row[col_idx]
        value = cell.value
        hyperlink = cell.hyperlink.target if cell.hyperlink else None
        return value, hyperlink

    def _get_cell_value(self, row, col_idx):
        """Get cell value."""
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx].value

    def _parse_eln_label(self, label):
        """
        Parse ELN label like 'KF001' into (initials, sequence_number).

        Returns (initials, seq_num) or (None, None) if invalid.
        """
        if not label:
            return None, None
        match = re.match(r'^([A-Za-z]{1,4})(\d+)$', str(label).strip())
        if match:
            initials = match.group(1).upper()
            seq_num = int(match.group(2))
            return initials, seq_num
        return None, None

    def _process_row(self, row, row_idx, column_map, target,
                     Supplier, Compound, Batch, LabNotebookEntry, CompoundDocument,
                     stats, suppliers_created, notebook_entries_created,
                     compounds_created, documents_created):
        """Process a single row from the spreadsheet."""

        # Get SMILES - required
        smiles = self._get_cell_value(row, column_map.get('smiles'))
        if not smiles:
            stats['skipped_empty'] += 1
            return None

        # Check for existing compound by SMILES
        if self.skip_existing:
            try:
                from rdkit import Chem
                mol = Chem.MolFromSmiles(smiles)
                if mol:
                    canonical = Chem.MolToSmiles(mol, canonical=True)
                    if Compound.objects.filter(rdkit_smiles=canonical).exists():
                        stats['skipped_existing'] += 1
                        self.log(f"Row {row_idx}: Skipping existing compound (SMILES match)")
                        return None
            except ImportError:
                pass

        # Get ELN reference (hyperlinked)
        eln_label, eln_url = self._get_cell_with_hyperlink(row, column_map.get('eln_reference'))
        initials, seq_num = self._parse_eln_label(eln_label)

        # Get or create supplier
        supplier = None
        notebook_entry = None
        if initials:
            supplier = Supplier.objects.filter(initials__iexact=initials).first()
            if not supplier:
                if self.create_suppliers:
                    # Try to get synthesized_by for the name
                    synth_by = self._get_cell_value(row, column_map.get('synthesized_by'))
                    name = synth_by if synth_by else f"Chemist {initials}"
                    supplier = Supplier.objects.create(
                        name=name,
                        initials=initials.upper()
                    )
                    suppliers_created.append(f"{supplier.name} ({supplier.initials})")
                    self.log(f"Row {row_idx}: Created supplier {supplier.name}")
                else:
                    self.stdout.write(self.style.WARNING(
                        f"Row {row_idx}: Unknown supplier initials '{initials}' - "
                        f"use --create-suppliers to auto-create"
                    ))

            # Get or create notebook entry
            if supplier and seq_num is not None:
                notebook_entry, created = LabNotebookEntry.objects.get_or_create(
                    supplier=supplier,
                    sequence_number=seq_num,
                    defaults={'url': eln_url or ''}
                )
                if created:
                    notebook_entries_created.append(notebook_entry.label)
                    self.log(f"Row {row_idx}: Created notebook entry {notebook_entry.label}")
                elif eln_url and not notebook_entry.url:
                    # Update URL if entry exists but had no URL
                    notebook_entry.url = eln_url
                    notebook_entry.save(update_fields=['url'])

        # Get other fields
        sequence = self._get_cell_value(row, column_map.get('sequence'))
        helm = self._get_cell_value(row, column_map.get('helm'))
        formula = self._get_cell_value(row, column_map.get('chemical_formula'))
        exact_mass = self._get_cell_value(row, column_map.get('exact_mass'))
        peptide_id = self._get_cell_value(row, column_map.get('peptide_id'))

        # Create compound
        compound = Compound(
            target=target,
            smiles=smiles,
            supplier=supplier,
            supplier_ref=peptide_id or '',
            notebook_entry=notebook_entry,
            sequence_display=sequence or '',
            helm_notation=helm or '',
            comments=f"Formula: {formula}" if formula else '',
        )

        # Set molecular weight from exact mass if available
        if exact_mass:
            try:
                compound.molecular_weight = float(exact_mass)
            except (ValueError, TypeError):
                pass

        compound.save()
        compounds_created.append(compound.formatted_id)
        self.log(f"Row {row_idx}: Created compound {compound.formatted_id}")

        # Create documents for hyperlinked files
        doc_columns = [
            ('chemdraw_file', 'chemdraw'),
            ('lcms', 'qc'),
            ('hrms', 'qc'),
        ]
        for col_name, doc_kind in doc_columns:
            label, url = self._get_cell_with_hyperlink(row, column_map.get(col_name))
            if url:
                doc = CompoundDocument.objects.create(
                    compound=compound,
                    kind=doc_kind,
                    label=str(label) if label else '',
                    url=url,
                )
                documents_created.append(f"{compound.formatted_id}/{doc.kind}")
                self.log(f"Row {row_idx}: Created {doc_kind} document")

        return compound
